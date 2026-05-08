import json
import os
from datetime import date

import openpyxl


SOURCE_XLSX = '鋁線槽系列規格表.xlsx'
SHEET_NAME = '鋁線槽系列規格表'
GALLERY_DIR = '圖庫'
OUTPUT_JSON = 'products.json'


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s == '' else s


def to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_price(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    return None if s == '' else s


def split_tags(v):
    s = clean(v)
    if not s or s == 'N/A':
        return (s, [])
    items = [x.strip() for x in s.split(',') if x.strip() and x.strip() != 'N/A']
    return (s, items)


def build_image_map(gallery_dir=GALLERY_DIR):
    image_map = {}
    if not os.path.isdir(gallery_dir):
        return image_map

    for sku_folder in os.listdir(gallery_dir):
        folder_path = os.path.join(gallery_dir, sku_folder)
        if not os.path.isdir(folder_path):
            continue

        files = [
            f for f in os.listdir(folder_path)
            if f.lower().endswith(('.jpg', '.jpeg', '.png'))
            and not f.startswith('.')
        ]
        if not files:
            continue

        main = None
        for priority in ['外觀1.jpeg', '外觀1.jpg', '外觀1.png']:
            if priority in files:
                main = priority
                break

        if not main:
            for f in sorted(files):
                if f.startswith('外觀'):
                    main = f
                    break

        if not main:
            main = sorted(files)[0]

        gallery = [
            f for f in sorted(files)
            if not f.startswith('全頁參考') and not f.startswith('提取_')
        ][:5]

        image_map[sku_folder] = {
            'main': f'{gallery_dir}/{sku_folder}/{main}',
            'gallery': [f'{gallery_dir}/{sku_folder}/{f}' for f in gallery],
        }

    return image_map


def padded(row_values, length=22):
    return list(row_values) + [None] * max(0, length - len(row_values))


def row_to_product(row_values, image_map):
    row_values = padded(row_values)
    sku = clean(row_values[3])
    if not sku:
        return None

    color_temp_raw, color_temp_list = split_tags(row_values[13])
    scenes_raw, scene_list = split_tags(row_values[15])
    img = image_map.get(sku, {})
    has_image = bool(img.get('main'))

    return {
        'sku': sku,
        'catalogPage': clean(row_values[1]),
        'category': clean(row_values[2]),
        'name': clean(row_values[4]),
        'widthMm': to_float(row_values[5]),
        'heightMm': to_float(row_values[6]),
        'channelWidthMm': to_float(row_values[7]),
        'thicknessMm': to_float(row_values[8]),
        'mountMethod': clean(row_values[9]),
        'mountLocation': clean(row_values[10]),
        'lightDirection': clean(row_values[11]),
        'voltage': clean(row_values[12]),
        'colorTemp': color_temp_raw,
        'colorTempList': color_temp_list,
        'ipRating': clean(row_values[14]),
        'scenes': scenes_raw,
        'sceneList': scene_list,
        'features': clean(row_values[16]),
        'colors': clean(row_values[17]),
        'relatedParts': clean(row_values[18]),
        'price': parse_price(row_values[19]),
        'priceUnit': clean(row_values[20]),
        'notes': clean(row_values[21]),
        'hasImage': has_image,
        'images': img if has_image else {'main': None, 'gallery': []},
    }


def main():
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
    image_map = build_image_map(GALLERY_DIR)

    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        product = row_to_product(row, image_map)
        if product:
            products.append(product)

    output = {
        'meta': {
            'generated': str(date.today()),
            'total': len(products),
            'source': SOURCE_XLSX,
        },
        'products': products,
    }

    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f'✓ 輸出 {OUTPUT_JSON}，共 {len(products)} 筆')


if __name__ == '__main__':
    main()
